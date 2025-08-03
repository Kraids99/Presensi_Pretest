from flask import Flask, render_template, request, send_file
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

app = Flask(__name__, static_folder='../static')

@app.route('/')
def index():
    return send_file('../index.html')

@app.route('/proses', methods=['POST'])
def proses():
    log_file = request.files['log']
    pretest_file = request.files['pretest']

    def read_file(input_file):
        filename = input_file.filename
        if filename.endswith('.csv'):
            return pd.read_csv(input_file)
        elif filename.endswith(('.xls', '.xlsx')):
            return pd.read_excel(input_file)
        else:
            raise ValueError("Format file tidak didukung.")

    try:
        df_log = read_file(log_file)
        df_pretest = read_file(pretest_file)
    except Exception as e:
        return f"Gagal membaca file: {e}", 400

    df_attempts = df_log[
        df_log["Event name"].str.contains("attempt started", case=False, na=False)
    ]

    def is_ip_luar(ip):
        return not (str(ip).startswith("10.31.211"))

    def catatan_ip(ip):
        return "Mengerjakan dari luar jaringan" if is_ip_luar(ip) else "-"

    df_log_result = df_attempts[["Time", "IP address", "User full name"]].copy()
    df_log_result["Status"] = "Hadir"
    df_log_result["Catatan"] = df_log_result["IP address"].apply(catatan_ip)

    df_pretest["User full name"] = df_pretest["First name"].astype(str).str.strip() + " " + df_pretest["Last name"].astype(str).str.strip()
    df_pretest["NPM"] = df_pretest["Email address"].str.split("@").str[0]
    df_pretest["Nilai"] = df_pretest["Grade/100.00"].str.split(".").str[0]

    df_pretest_result = df_pretest[["NPM", "User full name", "Nilai"]].iloc[:-1]

    df_merge_result = pd.merge(
        df_log_result,
        df_pretest_result,
        on="User full name",
        how="right"
    )

    df_merge_result = df_merge_result[["Time", "IP address", "NPM", "User full name", "Nilai", "Status", "Catatan"]]
    df_merge_result = df_merge_result.sort_values(by="NPM", ascending=True)

    df_merge_result["Status"] = df_merge_result["Status"].fillna("Tidak Hadir")
    df_merge_result["Catatan"] = df_merge_result["Catatan"].fillna("-")
    df_merge_result["Time"] = df_merge_result["Time"].fillna("-")
    df_merge_result["IP address"] = df_merge_result["IP address"].fillna("-")
    df_merge_result["Nilai"] = df_merge_result["Nilai"].fillna("-")

    # simpan ke memory (BytesIO)
    output = BytesIO()
    df_merge_result.to_excel(output, index=False)
    output.seek(0)

    # format Excel
    wb = load_workbook(output)
    ws = wb.active

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            cell.font = Font(name='Times New Roman', size=12)

            if i == 1:
                cell.fill = header_fill
                cell.font = Font(name='Times New Roman', size=14, bold=True)
            elif i % 2 == 0:
                cell.fill = alt_fill

    ws.auto_filter.ref = ws.dimensions

    output_final = BytesIO()
    wb.save(output_final)
    output_final.seek(0)

    return send_file(
        output_final,
        download_name="Hasil_Presensi.xlsx",
        as_attachment=True
    )

# if __name__ == '__main__':
#     app.run(debug=True)

if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)

# def handler(environ, start_response):
#     return app(environ, start_response)
