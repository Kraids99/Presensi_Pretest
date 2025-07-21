import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

root = tk.Tk()
root.title("Auto-Presensi") 
root.withdraw()

input_path_log = tk.filedialog.askopenfilename(
    title="Pilih file report log (CSV/Excel)",
    filetypes=[("All supported", "*.csv *.xlsx *.xls")]
)

input_path_pretest = tk.filedialog.askopenfilename(
    title="Pilih file report pretest (CSV/Excel)",
    filetypes=[("All supported", "*.csv *.xlsx *.xls")]
)

def read_file(input_path):
    if input_path.endswith('.csv'):
        return pd.read_csv(input_path)
    elif input_path.endswith(('.xls', '.xlsx')):
        return pd.read_excel(input_path)
    else:
        raise ValueError("Format file tidak didukung.")

if not (input_path_log and input_path_pretest):
    print("Tidak ada file yang dipilih. Program berhenti.")
    exit()

try:
    df_log = read_file(input_path_log)
    df_pretest = read_file(input_path_pretest)
except Exception as e:
    print(f"Gagal membaca file: {e}")
    exit()

print("File berhasil dibaca.")

df_attempts = df_log[
    df_log["Event name"].str.contains("attempt started", case=False, na=False)
]

def is_ip_luar(ip):
    return not (str(ip).startswith("10.31.211"))

def catatan_ip(ip):
    if(is_ip_luar(ip)):
        return "Mengerjakan dari luar jaringan"
    else:
        return "-"

df_log_result = df_attempts[["Time", "IP address", "User full name"]].copy()
df_log_result["Status"] = "Hadir"
df_log_result["Catatan"] = df_log_result["IP address"].apply(catatan_ip)

df_pretest["User full name"] = df_pretest["First name"].astype(str).str.strip() + " " + df_pretest["Last name"].astype(str).str.strip()
df_pretest["NPM"] = df_pretest["Email address"].str.split("@").str[0]
df_pretest["Nilai"] = df_pretest["Grade/100.00"].str.split(".").str[0]

df_pretest_result = df_pretest[["NPM", "User full name", "Nilai"]]
df_pretest_result = df_pretest.iloc[:-1]
# Cara lain
# df_pretest_result = df_pretest.drop(df_pretest.tail(1).index)

df_merge_result = pd.merge(
    df_log_result,
    df_pretest_result,
    on="User full name",
    how="right"  # tetap ada meskipun di log g ada
)

df_merge_result = df_merge_result[["Time", "IP address", "NPM", "User full name", "Nilai", "Status", "Catatan"]]

df_merge_result = df_merge_result.sort_values(by="NPM", ascending=True)


df_merge_result["Status"] = df_merge_result["Status"].fillna("Tidak Hadir")
df_merge_result["Catatan"] = df_merge_result["Catatan"].fillna("-")
df_merge_result["Time"] = df_merge_result["Time"].fillna("-")
df_merge_result["IP address"] = df_merge_result["IP address"].fillna("-")
df_merge_result["Nilai"] = df_merge_result["Nilai"].fillna("-")

# Debugging
print("Jumlah hasil presensi:", len(df_merge_result))
print(df_merge_result.head())

output_path = filedialog.asksaveasfilename(
    title="Simpan hasil presensi sebagai...",
    defaultextension=".xlsx",
    filetypes=[("Excel Files", "*.xlsx")],
    initialfile="Hasil_Presensi.xlsx"
)

if not output_path:
    print("Output file tidak dipilih. Program dihentikan.")
    exit()

try:
    df_merge_result.to_excel(output_path, index=False)
    # Excel
    wb = load_workbook(output_path)
    ws = wb.active

    # border dan warna
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Styling
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            cell.font = Font(name='Times New Roman', size=12)

            # Header
            if i == 1:
                cell.fill = header_fill
                cell.font = Font(name='Times New Roman', size=14, bold=True)
            elif i % 2 == 0:  # striped row
                cell.fill = alt_fill

    # Filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"Presensi berhasil diekspor ke: {output_path}")
except Exception as e:
    print(f"Gagal menyimpan file: {e}")