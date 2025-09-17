from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

def format_excel(output_path):
    # Buka workbook yang baru disimpan
    wb = load_workbook(output_path)
    ws = wb.active

    # Border dan warna
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    alt_fill   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Ganti nama sheet
    ws.title = "Presensi"

    # Atur border & styling
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column),start=1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            cell.font = Font(name='Times New Roman', size=12)

            # Header
            if i == 1:
                cell.fill = header_fill
                cell.font = Font(name='Times New Roman', size=14, bold=True)
            elif i % 2 == 0:
                cell.fill = alt_fill

        # Pewarnaan merah untuk baris "Tidak Hadir"
        status_cell = row[5]
        if status_cell.value == "Tidak Hadir":
            for c in row:
                c.fill = red_fill

    # Atur lebar kolom
    column_widths = [18, 18, 12, 30, 10, 15, 20]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Aktifkan filter
    ws.auto_filter.ref = ws.dimensions

    # Simpan perubahan
    wb.save(output_path)
