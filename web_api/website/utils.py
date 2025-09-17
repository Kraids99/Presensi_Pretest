import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

def read_file(input_file):
    name = input_file.filename.lower()
    if name.endswith('.csv'):
        return pd.read_csv(input_file)
    elif name.endswith(('.xls', '.xlsx')):
        return pd.read_excel(input_file)
    raise ValueError("Format file tidak didukung.")

def is_ip_luar(ip):
    return not (str(ip).startswith("10.31.211"))

def catatan_ip_list(ip_list):
    return "Mengerjakan dari luar jaringan" if any(is_ip_luar(ip) for ip in ip_list) else "-"

def pilih_ip_tampil(ip_list):
    """
    - Jika semua IP diawali '10.31.211' -> ambil IP pertama
    - Jika ada IP luar -> ambil salah satu IP luar (yang pertama ditemukan)
    """
    if not ip_list:
        return "-"
    # cek apakah ada ip luar
    for ip in ip_list:
        if is_ip_luar(ip):
            return ip  # ambil IP luar pertama
    # kalau semua internal
    return ip_list[0]

from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

def style_excel(df_merge_result):
    """
    Simpan DataFrame ke BytesIO Excel dengan styling:
    - Border tipis
    - Header biru, baris selang-seling abu
    - Baris merah untuk Status 'Tidak Hadir'
    - Lebar kolom dan auto filter
    """
    # simpan dataframe ke memori sementara
    output = BytesIO()
    df_merge_result.to_excel(output, index=False)
    output.seek(0)

    # buka workbook dari memori
    wb = load_workbook(output)
    ws = wb.active
    ws.title = "Presensi"

    # border & fill style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    alt_fill    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # loop semua cell untuk border, font, dan warna
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column),
        start=1
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center",
                                       wrap_text=True)
            cell.font = Font(name='Times New Roman', size=12)

            if i == 1:  # header
                cell.fill = header_fill
                cell.font = Font(name='Times New Roman', size=14, bold=True)
            elif i % 2 == 0:  # baris selang-seling
                cell.fill = alt_fill

        # pewarnaan merah jika Status == "Tidak Hadir"
        status_cell = row[5]  # kolom ke-6 (index 5) = Status
        if status_cell.value == "Tidak Hadir":
            for c in row:
                c.fill = red_fill

    # atur lebar kolom
    column_widths = [18, 18, 12, 30, 10, 15, 20]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # aktifkan filter otomatis
    ws.auto_filter.ref = ws.dimensions

    # simpan ke BytesIO untuk dikirim lewat Flask send_file
    output_final = BytesIO()
    wb.save(output_final)
    output_final.seek(0)
    return output_final
